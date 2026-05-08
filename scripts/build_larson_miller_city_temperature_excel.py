#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""流式生成 Larson-Miller 榜单关键城市温度 Excel。

artifact-tool 在 8 个城市长表的大工作簿导出阶段内存压力较大，本脚本使用
openpyxl 的 write_only 模式生成标准 .xlsx。
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


START_YEAR = 2016
END_YEAR = 2025
TIME_STANDARD = "lst"
CITY_IDS = [
    "ahvaz",
    "basra",
    "jacobabad",
    "riyadh",
    "mecca",
    "kuwait_city",
    "jizan_saudi",
    "port_sudan",
]
OUTPUT = Path("outputs/Larson-Miller榜单关键城市_NASA_POWER小时气温_2016_2025.xlsx")


def slug() -> str:
    return "_".join(CITY_IDS)


def csv_path_for_city(city_id: str) -> Path:
    return Path(f"data/processed/{city_id}_t2m_hourly_{START_YEAR}_{END_YEAR}_{TIME_STANDARD}.csv")


def wide_csv_path() -> Path:
    return Path(f"data/processed/nasa_power_t2m_by_date_hour_{START_YEAR}_{END_YEAR}_{slug()}_{TIME_STANDARD}.csv")


def summary_json_path() -> Path:
    return Path(f"data/summary/nasa_power_t2m_export_summary_{START_YEAR}_{END_YEAR}_{slug()}_{TIME_STANDARD}.json")


def append_csv_sheet(wb: Workbook, title: str, path: Path) -> int:
    ws = wb.create_sheet(title=title)
    count = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            ws.append(row)
            count += 1
    return count


def set_basic_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_summary_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet(title="摘要")
    rows = [
        ["项目", "内容"],
        ["数据源", summary["source"]],
        ["接口", summary["endpoint"]],
        ["参数", f'{summary["parameter"]} ({summary["parameter_meaning"]})'],
        ["时间标准", f'{summary["time_standard"]} (当地太阳时)'],
        ["年份范围", f'{summary["start_year"]}-{summary["end_year"]}'],
        ["城市数量", len(CITY_IDS)],
        ["合并长表行数", summary["combined_rows"]],
        ["宽表行数", summary["wide_rows"]],
        ["说明", "T2M 为 2 米气温小时平均值，单位摄氏度；宽表按 date + hour_lst 对齐全部城市。"],
        [],
        ["city_id", "城市", "国家", "纬度", "经度", "小时数", "缺失值", "温度范围", "平均温度"],
    ]
    for city_id in CITY_IDS:
        city = summary["cities"][city_id]
        rows.append(
            [
                city_id,
                f'{city["city_zh"]} / {city["city_en"]}',
                f'{city["country_zh"]} / {city["country_en"]}',
                city["latitude"],
                city["longitude"],
                city["row_count"],
                city["missing_count"],
                f'{city["t2m_c_min"]} 至 {city["t2m_c_max"]} °C',
                f'{city["t2m_c_mean"]} °C',
            ]
        )

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for cell in ws[12]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    set_basic_widths(
        ws,
        {
            "A": 18,
            "B": 30,
            "C": 24,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 10,
            "H": 20,
            "I": 16,
        },
    )


def verify_workbook(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        row_counts = {name: wb[name].max_row for name in sheets}
        return {"sheets": sheets, "row_counts": row_counts}
    finally:
        wb.close()


def main() -> int:
    summary = json.loads(summary_json_path().read_text(encoding="utf-8"))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=False)
    default = wb.active
    wb.remove(default)

    build_summary_sheet(wb, summary)
    append_csv_sheet(wb, "宽表_全部城市对齐", wide_csv_path())

    for city_id in CITY_IDS:
        city = summary["cities"][city_id]
        sheet_name = f'{city["city_zh"]}_长表'
        append_csv_sheet(wb, sheet_name, csv_path_for_city(city_id))

    wb.save(OUTPUT)
    check = verify_workbook(OUTPUT)
    print(json.dumps(check, ensure_ascii=False, indent=2))
    print(f"saved {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

